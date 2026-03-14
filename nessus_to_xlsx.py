"""
Nessus to Excel (XLSX) Conversion Script

This script converts .nessus (XML) vulnerability scan files into Microsoft Excel
format (.xlsx) with formatted tables for easy analysis. It processes .nessus files
and extracts comprehensive vulnerability data, then creates a formatted Excel
workbook with:

- Formatted table with alternating row colors for readability
- All vulnerability data in structured columns
- Host information, vulnerability details, CVSS scores, CVE IDs
- Exploit information and security references

Features:
- Creates formatted Excel tables with professional styling
- Supports processing multiple .nessus files
- Removes duplicate findings automatically
- Provides progress tracking and performance metrics

Note: Variable names (e.g., pluginName, Credentialed_Scan) mirror the Nessus XML
element and attribute names for traceability back to the source format.
"""

# Import required libraries.
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
import tkinter as tk
from tkinter import filedialog
import psutil
from datetime import datetime

# Severity constants (defined once, not per iteration)
SEVERITY_MAPPING = {0: 4, 1: 3, 2: 2, 3: 1, 4: 0}
SEVERITY_RATING = {0: "Critical", 1: "High", 2: "Medium", 3: "Low", 4: "Informational"}


def get_text(parent, xpath, default=""):
    """Extract text from an XML element by XPath, returning default if not found."""
    el = parent.find(xpath)
    return el.text if el is not None else default


# Create a file dialog window for the user to select multiple XML files.
root = tk.Tk()
root.withdraw()
xml_files = filedialog.askopenfilenames(title="Select files to process",filetypes=[("Nessus files", "*.nessus")])

# Start run timer.
start_time = time.time()

# Create an empty list to store the data.
data = []

# Loop through each XML file.
for xml_file in xml_files:
    # Extract just the file name from the file path
    file_name = os.path.basename(xml_file)

    # Parse the XML file.
    try:
        tree = ET.parse(xml_file)
    except ET.ParseError as e:
        print(f"Error parsing {xml_file}: {e}")
        continue

    xml_root = tree.getroot()
    print(f"Processing {xml_file}")

    # Get the total number of hosts.
    total_hosts = len(list(xml_root.iter("ReportHost")))

    # Process each host.
    for i, host in enumerate(xml_root.iter("ReportHost"), 1):
        host_cpe = get_text(host, "HostProperties/tag[@name='cpe']")
        Credentialed_Scan = get_text(host, "HostProperties/tag[@name='Credentialed_Scan']")
        host_end = get_text(host, "HostProperties/tag[@name='HOST_END']")
        host_end_timestamp = get_text(host, "HostProperties/tag[@name='HOST_END_TIMESTAMP']")
        host_start = get_text(host, "HostProperties/tag[@name='HOST_START']")
        host_start_timestamp = get_text(host, "HostProperties/tag[@name='HOST_START_TIMESTAMP']")
        host_fqdn = get_text(host, "HostProperties/tag[@name='host-fqdn']")
        host_ip = get_text(host, "HostProperties/tag[@name='host-ip']")
        hostname = get_text(host, "HostProperties/tag[@name='hostname']")
        host_rdns = get_text(host, "HostProperties/tag[@name='host-rdns']")
        LastAuthenticatedResults = get_text(host, "HostProperties/tag[@name='LastAuthenticatedResults']")
        local_checks_proto = get_text(host, "HostProperties/tag[@name='local-checks-proto']")
        mac_address = get_text(host, "HostProperties/tag[@name='mac-address']")
        netbios_name = get_text(host, "HostProperties/tag[@name='netbios-name']")
        operating_system = get_text(host, "HostProperties/tag[@name='operating-system']")
        operating_system_conf = get_text(host, "HostProperties/tag[@name='operating-system-conf']")
        operating_system_method = get_text(host, "HostProperties/tag[@name='operating-system-method']")
        operating_system_id = get_text(host, "HostProperties/tag[@name='os']")
        patch_summary_total_cves = get_text(host, "HostProperties/tag[@name='patch-summary-total-cves']")
        policy_used = get_text(host, "HostProperties/tag[@name='policy-used']")
        sinfp_signature = get_text(host, "HostProperties/tag[@name='sinfp-signature']")
        smb_login_used = get_text(host, "HostProperties/tag[@name='smb-login-used']")
        system_type = get_text(host, "HostProperties/tag[@name='system-type']")
        virtual_mac_address = get_text(host, "HostProperties/tag[@name='virtual-mac-address']")
        wmi_domain = get_text(host, "HostProperties/tag[@name='wmi-domain']")

        # Extract scan_date from host_start
        scan_date = None
        host_start_tag = host.find('HostProperties/tag[@name="HOST_START"]')
        if host_start_tag is not None:
            host_start = host_start_tag.text
            if host_start:
                try:
                    scan_date = datetime.strptime(host_start, "%a %b %d %H:%M:%S %Y").strftime("%Y-%m-%d")
                except ValueError:
                    pass

        # Loop through each report item
        for item in host.iter("ReportItem"):
            pluginFamily = item.attrib["pluginFamily"]
            pluginID = item.attrib["pluginID"]
            port = item.attrib["port"]
            protocol = item.attrib["protocol"]
            severity = item.attrib["severity"]

            severity_reversed = SEVERITY_MAPPING[int(severity)]
            severity_rating = SEVERITY_RATING[severity_reversed]

            svc_name = item.attrib["svc_name"]
            age_of_vuln = get_text(item, "age_of_vuln")
            agent = get_text(item, "agent")
            always_run = get_text(item, "always_run")
            asset_categories = get_text(item, "asset_categories")
            asset_inventory = get_text(item, "asset_inventory")
            asset_inventory_category = get_text(item, "asset_inventory_category")
            bid = get_text(item, "bid")
            canvas_package = get_text(item, "canvas_package")
            cea_id = get_text(item, "cea-id")
            cert = get_text(item, "cert")
            cisa_known_exploited = get_text(item, "cisa-known-exploited")
            cisa_ncas = get_text(item, "cisa-ncas")
            cisco_bug_id = get_text(item, "cisco-bug-id")
            cisco_sa = get_text(item, "cisco-sa")
            cpe = get_text(item, "cpe")
            cve_elements = item.findall('cve')
            cve = ', '.join(c.text for c in cve_elements) if cve_elements else ''
            cvss_base_score = get_text(item, "cvss_base_score")
            cvss_score_rationale = get_text(item, "cvss_score_rationale")
            cvss_score_source = get_text(item, "cvss_score_source")
            cvss_temporal_score = get_text(item, "cvss_temporal_score")
            cvss_temporal_vector = get_text(item, "cvss_temporal_vector")
            cvss_vector = get_text(item, "cvss_vector")
            cvss3_base_score = get_text(item, "cvss3_base_score")
            cvss3_score_source = get_text(item, "cvss3_score_source")
            cvss3_temporal_score = get_text(item, "cvss3_temporal_score")
            cvss3_temporal_vector = get_text(item, "cvss3_temporal_vector")
            cvss3_vector = get_text(item, "cvss3_vector")
            cvssV3_impactScore = get_text(item, "cvssV3_impactScore")
            cwe = get_text(item, "cwe")
            description = get_text(item, "description")
            edb_id = get_text(item, "edb-id")
            exploit_available = get_text(item, "exploit_available")
            exploit_code_maturity = get_text(item, "exploit_code_maturity")
            exploit_framework_canvas = get_text(item, "exploit_framework_canvas")
            exploit_framework_core = get_text(item, "exploit_framework_core")
            exploit_framework_metasploit = get_text(item, "exploit_framework_metasploit")
            exploitability_ease = get_text(item, "exploitability_ease")
            exploited_by_malware = get_text(item, "exploited_by_malware")
            exploited_by_nessus = get_text(item, "exploited_by_nessus")
            generated_plugin = get_text(item, "generated_plugin")
            hardware_inventory = get_text(item, "hardware_inventory")
            iava = get_text(item, "iava")
            iavb = get_text(item, "iavb")
            iavt = get_text(item, "iavt")
            icsa = get_text(item, "icsa")
            in_the_news = get_text(item, "in_the_news")
            metasploit_name = get_text(item, "metasploit_name")
            msft = get_text(item, "msft")
            mskb = get_text(item, "mskb")
            os_identification = get_text(item, "os_identification")
            patch_publication_date = get_text(item, "patch_publication_date")
            plugin_modification_date = get_text(item, "plugin_modification_date")
            pluginName = item.attrib.get("pluginName")
            plugin_output_element = item.find('plugin_output')
            plugin_output = plugin_output_element.text.strip() if plugin_output_element is not None and plugin_output_element.text else ""
            plugin_publication_date = get_text(item, "plugin_publication_date")
            plugin_type = get_text(item, "plugin_type")
            product_coverage = get_text(item, "product_coverage")
            risk_factor = get_text(item, "risk_factor")
            script_version = get_text(item, "script_version")
            secunia = get_text(item, "secunia")
            see_also = get_text(item, "see_also")
            solution = get_text(item, "solution")
            stig_severity = get_text(item, "stig_severity")
            synopsis = get_text(item, "synopsis")
            thorough_tests = get_text(item, "thorough_tests")
            threat_intensity_last_28 = get_text(item, "threat_intensity_last_28")
            threat_recency = get_text(item, "threat_recency")
            threat_sources_last_28 = get_text(item, "threat_sources_last_28")
            tra = get_text(item, "tra")
            unsupported_by_vendor = get_text(item, "unsupported_by_vendor")
            vmsa = get_text(item, "vmsa")
            vpr_score = get_text(item, "vpr_score")
            vuln_publication_date = get_text(item, "vuln_publication_date")

            # Column order: 106 fields matching XLSX headers
            data.append([host_ip, mac_address, virtual_mac_address, hostname, netbios_name, host_fqdn, host_rdns, system_type, cpe, operating_system_id, operating_system, severity_rating, risk_factor, severity_reversed, pluginName, synopsis, description, solution, port, protocol, svc_name, plugin_output, plugin_type, plugin_publication_date, plugin_modification_date, stig_severity, cvss_base_score, cvss_vector, cvss_score_rationale, cvss_score_source, cvss_temporal_score, cvss_temporal_vector, cvss3_base_score, cvss3_vector, cvss3_score_source, cvss3_temporal_score, cvss3_temporal_vector, cvssV3_impactScore, vpr_score, patch_publication_date, patch_summary_total_cves, vuln_publication_date, age_of_vuln, product_coverage, exploit_available, exploitability_ease, exploit_code_maturity, exploit_framework_canvas, canvas_package, exploit_framework_core, exploit_framework_metasploit, metasploit_name, exploited_by_malware, exploited_by_nessus, edb_id, threat_recency, threat_intensity_last_28, threat_sources_last_28, local_checks_proto, smb_login_used, wmi_domain, thorough_tests, Credentialed_Scan, LastAuthenticatedResults, policy_used, host_start, host_start_timestamp, host_end, host_end_timestamp, os_identification, operating_system_conf, operating_system_method, sinfp_signature, pluginID, pluginFamily, script_version, agent, always_run, asset_inventory, asset_inventory_category, asset_categories, hardware_inventory, bid, cea_id, cert, cisa_known_exploited, cisa_ncas, cisco_bug_id, cisco_sa, cve, cwe, iava, iavb, iavt, icsa, msft, mskb, tra, vmsa, secunia, unsupported_by_vendor, see_also, in_the_news, generated_plugin, file_name, scan_date])

# Define remove duplicates function.
def remove_duplicates(data):
    """
    Removes duplicate rows from the extracted vulnerability data.

    Args:
        data: List of data rows (each row is a list of values)

    Returns:
        A tuple containing:
        - List of unique rows (duplicates removed)
        - Count of duplicates that were removed
    """
    unique_rows = set()
    duplicates_removed = 0
    for row in data:
        row_tuple = tuple(row)
        if row_tuple not in unique_rows:
            unique_rows.add(row_tuple)
        else:
            duplicates_removed += 1
    return [list(row_tuple) for row_tuple in unique_rows], duplicates_removed

# Remove duplicates from the data.
data, duplicates_removed = remove_duplicates(data)

# Print a message displaying the number of removed duplicates.
if duplicates_removed > 0:
    print(f"Removed {duplicates_removed} duplicate rows from the data.")
else:
    print("No duplicate rows found.")

# Define the default filename
xlsx_filename = "nessus_conversion.xlsx"

# Open a file dialog to choose the output filename and directory
xlsx_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=xlsx_filename, title="Save XLSX File As...")

# Write the data to an XLSX file with a progress indicator.
print(f"Processed {len(xml_files)} files, {total_hosts} hosts, and {len(data)} findings.\nNow writing to XLSX: {xlsx_filename}.")
wb = Workbook()
ws = wb.active
assert ws is not None  # New Workbook always has an active worksheet
ws.append(["IP Address", "MAC Address", "Virtual MAC Address", "Hostname", "NETBIOS", "FQDN", "RDNS", "System Type", "CPE", "OS Family", "Operating System", "Severity Rating", "Risk Factor", "Severity Number", "Plugin Name", "Synopsis", "Description", "Solution", "Port", "Protocol", "Service Name", "Plugin Output", "Plugin Type", "Plugin Publication Date", "Plugin Modification Date", "STIG Severity", "CVSS Score", "CVSS Vector", "CVSS Score Rationale", "CVSS Score Source", "CVSS Temporal Score", "CVSS Temporal Vector", "CVSS3 Score", "CVSS3 Vector", "CVSS3 Score Source", "CVSS3 Temporal Score", "CVSS3 Temporal Vector", "CVSS3 Impact Score", "VPR Score", "Patch Publication Date", "Patch Summary CVE Count", "Vulnerability Publication Date", "Vulnerability Age", "Product Coverage", "Exploit Available", "Exploitability Ease", "Exploit Code Maturity", "Exploitable with CANVAS", "CANVAS Package", "Exploitable with Core Impact", "Exploitable with Metasploit", "Metasploit Module", "Exploitable by Malware", "Exploitable by Nessus", "Exploit DB ID", "Threat Recency", "Threat Intensity Last 28 Days", "Threat Sources Last 28 Days", "Local Checks Protocol", "SMB Login Used", "WMI Domain", "Thorough Tests Used", "Credentialed Scan Used", "Authentication Issues", "Policy Used", "Host Scan Started", "Host Scan Started Timestamp", "Host Scan Ended", "Host Scan Ended Timestamp", "OS Identified", "Operating System Confidence", "OS Identification Method", "SinFP Signature", "Plugin ID", "Plugin Family", "Script Version", "Agent Used", "Always Run", "Asset Inventory", "Asset Inventory Category", "Asset Categories", "Hardware Inventory", "BugtraqID", "CEA Id", "CERT", "CISA Known Exploited", "CISA NCAS", "Cisco Bug ID", "Cisco Security Advisory", "CVE", "CWE", "IAVA", "IAVB", "IAVT", "ICSA", "MSFT", "MSKB", "Trust Research Advisory", "VMWare Security Advisory", "Secunia Security Advisory", "Unsupported by Vendor", "See Also", "In The News", "Generated Plugin", "File Name", "Scan Date"])
for progress, row in enumerate(data, start=1):
    percent_complete = (progress / len(data)) * 100
    print(f"\rWriting data to XLSX: {percent_complete:.2f}% complete.", end="", flush=True)
    ws.append(row)
print("\rWriting data to XLSX: 100.00% complete.")

# Define the table range and create the table
print("Converting XLSX data to table")
table_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
table = Table(displayName="Table1", ref=table_range)

# Define the table style
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

# Apply the table style and save the file
table.tableStyleInfo = style
ws.add_table(table)
wb.save(xlsx_filename)

# Confirm completion and provide elapsed run time, peak CPU usage and peak memory usage.
print("Processing complete.")
end_time = time.time()
elapsed_time = end_time - start_time
peak_cpu_usage = psutil.Process().cpu_percent()
mem_info = psutil.Process().memory_info()
peak_mem_usage = getattr(mem_info, 'peak_wset', mem_info.rss) / 1024 / 1024
print(f"Elapsed time: {elapsed_time:.2f} seconds. Peak CPU usage: {peak_cpu_usage}%. Peak memory usage: {peak_mem_usage:.2f} MB.")
